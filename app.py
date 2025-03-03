# Importando as bibliotecas necessárias
import customtkinter as ctk  # Biblioteca para criar interfaces gráficas modernas
from tkinter import *  # Biblioteca padrão do Python para interfaces gráficas
from tkinter import messagebox  # Para exibir caixas de mensagem
import openpyxl  # Para manipular arquivos Excel
import pathlib  # Para manipular caminhos de arquivos

from customtkinter import CTkTextbox  # Componente de caixa de texto do CustomTkinter
from openpyxl import Workbook  # Para criar e manipular planilhas Excel

# Configuração da aparência padrão do CustomTkinter
ctk.set_appearance_mode("System")  # Define o modo de aparência (Light, Dark, System)
ctk.set_default_color_theme("blue")  # Define o tema de cores padrão


# Classe principal do aplicativo
class App(ctk.CTk):
    def __init__(self):
        super().__init__()  # Inicializa a classe pai (CTk)
        self.leayout_config()  # Configura o layout da janela
        self.apperance()  # Configura o botão de alternância de tema
        self.todo_sistema()  # Configura todo o sistema (interface e funcionalidades)

    # Configuração do layout da janela principal
    def leayout_config(self):
        self.title("Sistema de Gestão de Clientes")  # Título da janela
        self.geometry("700x550")  # Aumentei a altura para acomodar o novo campo

    # Configuração do botão de alternância de tema
    def apperance(self):
        self.lb_apm = ctk.CTkLabel(
            self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]
        )
        self.lb_apm.place(x=50, y=480)  # Ajustei a posição do rótulo "Tema"
        self.opt_atm = ctk.CTkOptionMenu(
            self, values=["Light", "Dark", "System"], command=self.change_apm
        )
        self.opt_atm.place(x=50, y=510)  # Ajustei a posição do menu de opções de tema

    # Configuração de todo o sistema (interface e funcionalidades)
    def todo_sistema(self):
        # Cria um frame para o título
        frame = ctk.CTkFrame(
            self,
            width=700,
            height=50,
            corner_radius=0,
            bg_color="teal",
            fg_color="teal",
        )
        frame.place(x=0, y=10)  # Posiciona o frame
        title = ctk.CTkLabel(
            frame,
            text="Sistema Gestão de Clientes",
            font=("Century Gothic", 24, "bold"),
            text_color="#fff",
        )
        title.place(x=190, y=10)  # Posiciona o título

        # Rótulo de instrução para o usuário
        span = ctk.CTkLabel(
            self,
            text="Por favor, preencha todos os campos do formulário",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )
        span.place(x=50, y=70)

        # Verifica se o arquivo Excel já existe, caso contrário, cria um novo
        ficheiro = pathlib.Path("Clientes.xlsx")
        if ficheiro.exists():
            pass  # Se o arquivo já existe, não faz nada
        else:
            ficheiro = Workbook()  # Cria um novo arquivo Excel
            folha = ficheiro.active  # Seleciona a planilha ativa
            # Define os cabeçalhos das colunas
            folha["A1"] = "Nome"
            folha["B1"] = "Contato"
            folha["C1"] = "Idade"
            folha["D1"] = "Gênero"
            folha["E1"] = "Endereço"
            folha["F1"] = "Orçamento"
            folha["G1"] = "Observações"
            ficheiro.save("Clientes.xlsx")  # Salva o arquivo

        # Função para enviar os dados do formulário
        def submit():
            # Obtém os valores dos campos
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            budget = budget_value.get()
            obs = obs_entry.get(0.0, END)

            # Verifica se todos os campos foram preenchidos
            if (
                name == ""
                or contact == ""
                or age == ""
                or gender == ""
                or address == ""
                or budget == ""
            ):
                messagebox.showerror(
                    "Sistema", "ERRO!\nPor favor preencha todos campos!"
                )
            else:
                # Abre o arquivo Excel e adiciona os dados
                ficheiro = openpyxl.load_workbook("Clientes.xlsx")
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row + 1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contact)
                folha.cell(column=3, row=folha.max_row, value=age)
                folha.cell(column=4, row=folha.max_row, value=gender)
                folha.cell(column=5, row=folha.max_row, value=address)
                folha.cell(column=6, row=folha.max_row, value=budget)
                folha.cell(column=7, row=folha.max_row, value=obs)
                ficheiro.save(r"Clientes.xlsx")  # Salva o arquivo
                messagebox.showinfo(
                    "Sistema", "Dados salvos com sucesso"
                )  # Exibe mensagem de sucesso

        # Função para limpar todos os campos do formulário
        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            budget_value.set("")
            obs_entry.delete(0.0, END)

        # Variáveis para armazenar os valores dos campos
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        budget_value = StringVar()  # Nova variável para o campo "Orçamento"

        # Campos de entrada de texto
        name_entry = ctk.CTkEntry(
            self,
            textvariable=name_value,
            width=350,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        contact_entry = ctk.CTkEntry(
            self,
            textvariable=contact_value,
            width=200,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        age_entry = ctk.CTkEntry(
            self,
            textvariable=age_value,
            width=150,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        addres_entry = ctk.CTkEntry(
            self,
            textvariable=address_value,
            width=200,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )
        budget_entry = ctk.CTkEntry(
            self,
            textvariable=budget_value,
            width=200,
            font=("Century Gothic", 16),
            fg_color="transparent",
        )  # Novo campo "Orçamento"

        # Combobox para seleção de gênero
        gender_combobox = ctk.CTkComboBox(
            self,
            values=["Masculino", "Feminino"],
            font=("Century Gothic", 14),
            width=150,
        )
        gender_combobox.set(value="Masculino")  # Define o valor padrão

        # Caixa de texto para observações
        obs_entry = CTkTextbox(
            self,
            width=350,
            height=150,
            font=("Arial", 18),
            border_color="#aaa",
            border_width=2,
            fg_color="transparent",
        )

        # Rótulos para os campos
        lb_name = ctk.CTkLabel(
            self,
            text="Nome completo:",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )
        lb_contact = ctk.CTkLabel(
            self,
            text="Contato:",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )
        lb_age = ctk.CTkLabel(
            self,
            text="Idade:",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )
        lb_gender = ctk.CTkLabel(
            self,
            text="Gênero",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )
        lb_addrress = ctk.CTkLabel(
            self,
            text="Endereço:",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )
        lb_budget = ctk.CTkLabel(
            self,
            text="Orçamento:",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )  # Novo rótulo "Orçamento"
        lb_obs = ctk.CTkLabel(
            self,
            text="Observação:",
            font=("Century Gothic", 16, "bold"),
            text_color=["#000", "#fff"],
        )

        # Botões de ação
        btn_submit = ctk.CTkButton(
            self,
            text="Salvar dados".upper(),
            command=submit,
            fg_color="#151",
            hover_color="#131",
        ).place(x=300, y=470)
        btn_clear = ctk.CTkButton(
            self,
            text="Limpar campos".upper(),
            command=clear,
            fg_color="#555",
            hover_color="#333",
        ).place(x=500, y=470)

        # Posicionamento dos elementos na janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_addrress.place(x=50, y=190)
        addres_entry.place(x=50, y=220)

        lb_budget.place(x=450, y=260)  # Posiciona o rótulo "Orçamento"
        budget_entry.place(x=450, y=290)  # Posiciona o campo "Orçamento"

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=50, y=290)

    # Função para alternar o tema da interface
    def change_apm(self, new_apparence_mod):
        ctk.set_appearance_mode(new_apparence_mod)  # Altera o modo de aparência


# Inicialização da aplicação
if __name__ == "__main__":
    app = App()  # Cria uma instância da classe App
    app.mainloop()  # Inicia o loop principal da interface gráfica
